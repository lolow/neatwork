import json
import openpyxl

def convert_json_to_excel(json_file, excel_file):
    # Ouvrir le fichier JSON
    with open(json_file, 'r') as f:
        data = json.load(f)
    
    # Créer un nouveau classeur Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row = 1
    
    # Parcourir les clés du JSON
    for _, (key, value) in enumerate(data.items(), start=1):
        # Écrire la clé dans la première colonne
        sheet.cell(row=row, column=1, value=key)

        row += 1
        
        # Si la valeur est une matrice, écrire chaque ligne sur une ligne différente
        if key == "S":
            for row_index, row_values in enumerate(value):
                for col_index, val in enumerate(row_values, start=1):
                    sheet.cell(row=row, column=col_index, value=val)
                row += 1
        else:
            # Si la valeur est une liste, écrire chaque élément sur la ligne suivante
            if isinstance(value, list):
                for col_index, item in enumerate(value, start=1):
                    sheet.cell(row=row, column=col_index, value=item)
            else:
                # Si la valeur est un scalaire, écrire dans la ligne suivante
                sheet.cell(row=row, column=1, value=value)
        
        # Passe à la ligne suivante
        row += 2
    
    # Sauvegarder le classeur Excel
    workbook.save(excel_file)
    print("Conversion terminée avec succès.")

# Utilisation de la fonction pour convertir le fichier JSON en Excel
convert_json_to_excel("SimuJava/src/dat.json", "SimuJava/src/dat.xlsx")
